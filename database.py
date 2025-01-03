import sqlite3
from datetime import date, datetime
import openpyxl as ox
import shutil
import cv2

conn = sqlite3.connect('AllPersons.db')
curr = conn.cursor()
curr.execute("""CREATE TABLE IF NOT EXISTS consumer(
    userid INT PRIMARY KEY,
    fname TEXT,
    lname TEXT,
    post TEXT,
    file TEXT,
    status TEXT
    )""")
curr.execute("""CREATE TABLE IF NOT EXISTS time_tracking(
    userid INT,
    dates DATE,
    times TIME, 
    status TEXT
    )""")  # status - больничный (ПР, но где-то нужно это объяснить), прогул (ПР), работал (Я)
curr.execute("""CREATE TABLE IF NOT EXISTS schedule(
    userid INT,
    dates DATE,
    time_b TIME,
    time_e TIME
    )""")
conn.commit()

haar_face_cascade = cv2.CascadeClassifier("haarcascade_frontalface_default.xml")  # Загрузка каскадного классификатора


def photo(img, userID, count):
    scale_factor = 1.1  # коэфицент увеличения размера окна поиска на каждой итерации
    min_neighbords = 6  # размер окна
    gray_img = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    faces = haar_face_cascade.detectMultiScale(gray_img, scale_factor, min_neighbords)  # Поиск всех лиц
    for (x, y, w, h) in faces:
        f = cv2.resize(gray_img[y:y + h, x:x + w], (200, 200))  # Создание кадра для идентицикатора
        cv2.imwrite('db/identify/%s_%s.jpg' % (userID, str(count)), f)  # Запись этого кадра


if __name__ == "__main__":
    command = ""
    while command != 'end':
        command = input()
        user = ''
        if command == 'insert':  # СОЗДАНИЕ ПОЛЬЗОВАТЕЛЯ В ТАБЛИЦЕ
            print("Enter userID, fname, lname, post")
            user = input()
            user = user.split()
            user.append('db/identify/%s' % user[0])
            user.append('Работает')
            curr.execute("INSERT INTO consumer VALUES(?,?,?,?,?,?)", user)
            conn.commit()
        if command == 'delete':  # УДАЛЕНИЕ ПОЛЬЗОВАТЕЛЯ ИЗ ТАБЛИЦЫ
            print("Enter userID")
            userID = input()
            try:
                curr.execute("SELECT userid FROM consumer WHERE userid = ?", userID)
            except:
                print('userID not found')
                continue
            curr.execute("DELETE FROM consumer WHERE consumer.userid = ?", userID)
            conn.commit()
        if command == 'status':  # ОБНОВИТЬ СТАТУС РАБОТНИКА
            print("Enter userID")
            userID = input()
            try:
                curr.execute("SELECT userid FROM consumer WHERE userid = ?", userID)
            except:
                print('userID not found')
                continue
            print("Enter status")
            user_status = input()
            curr.execute("UPDATE consumer SET status = ? WHERE consumer.userid = ?",
                         (user_status, userID))
            conn.commit()
        if command == 'tabel':
            print("Enter userID")
            userID = input()
            try:
                curr.execute("SELECT userid FROM consumer WHERE userid = ?", userID)
            except:
                print('userID not found')
                continue
            print("Enter type of work")
            type = input()
            shutil.copyfile('Tabel.xlsx',
                            'db/Tabels/%sTabel.xlsx' % str(userID))  # CКОПИРОВАЛИ НЕЗАПОЛНЕННЫЙ ТАБЕЛЬ В ПАПКУ ЮЗЕРА
            wb = ox.load_workbook('db/Tabels/%sTabel.xlsx' % str(userID))
            # НОМЕР ТАБЕЛЯ
            wb['стр.1'].cell(1, 79).value = userID
            # ПЕРИОД
            wb['стр.1'].cell(4, 70).value = '30'
            wb['стр.1'].cell(4, 77).value = 'Ноября'
            wb['стр.1'].cell(4, 100).value = '22'
            # УЧРЕЖДЕНИЕ
            wb['стр.1'].cell(5, 25).value = 'МГТУ им. Баумана'
            # СТРУКТУРНОЕ ПОДРАЗДЕЛЕНИЕ
            wb['стр.1'].cell(6, 25).value = 'ИУ8-73'
            # ВИД ТАБЕЛЯ
            wb['стр.1'].cell(7, 25).value = 'Первичный'
            # СТАТИСТИКА РАБОЧИХ ДНЕЙ ИЗ БД
            if type == 'жесткий':
                curr.execute("""select dates, status from time_tracking where userid = ?""", userID)
                statistic_days = curr.fetchall()
                half_day_cof = 0
                count_working_day = 0
                for i in statistic_days:
                    if i[0] == 16:
                        half_day_cof = 7
                        wb['стр.1'].cell(13, 94).value = count_working_day
                    if i[1] == 'Я':
                        count_working_day += 1
                    if i[1] == 'Я':
                        wb['стр.1'].cell(13, 34 + (int(i[0]) - 1) * 4 + half_day_cof).value = i[1]
                    else:
                        wb['стр.1'].cell(13, 34 + (int(i[0]) - 1) * 4 + half_day_cof).value = i[1]
                wb['стр.1'].cell(13, 165).value = count_working_day
            elif type == 'свободный':
                curr.execute("""select dates, status from time_tracking where userid = ?""", userID)
                statistic_days = curr.fetchall()
                curr.execute("""select dates, time_b, time_e from schedule where userid = ?""", userID)
                schedule_days = curr.fetchall()
                half_day_cof = 0
                count_day_hours = 0
                count_day_minuts = 0
                for i in statistic_days:
                    day, begin_day, end_day = schedule_days[i[0] - 1]
                    end_day = datetime.strptime(end_day, '%H:%M:%S')
                    begin_day = datetime.strptime(begin_day, '%H:%M:%S')
                    end_day_hours = end_day.hour
                    end_day_minuts = end_day.minute
                    begin_day_hours = begin_day.hour
                    begin_day_minuts = begin_day.minute
                    if i[0] == 16:
                        half_day_cof = 7
                        if count_day_minuts == 0:
                            wb['стр.1'].cell(13, 94).value = str(count_day_hours)
                        else:
                            wb['стр.1'].cell(13, 94).value = str(count_day_hours) + ':' + str(count_day_minuts)
                    if i[1] == 'Я':
                        if end_day_minuts - begin_day_minuts < 0:
                            end_day_hours -= 1
                            end_day_minuts += 60
                            end_day_minuts -= begin_day_minuts
                        else:
                            end_day_minuts -= begin_day_minuts
                        end_day_hours -= begin_day_hours
                        count_day_hours += end_day_hours
                        if count_day_minuts + end_day_minuts > 60:
                            count_day_hours += 1
                            count_day_minuts += 60
                            count_day_minuts -= end_day_minuts
                    if i[1] == 'Я':
                        if end_day_minuts == 0:
                            wb['стр.1'].cell(13, 34 + (int(i[0]) - 1) * 4 + half_day_cof).value = str(
                                end_day_hours)
                        else:
                            wb['стр.1'].cell(13, 34 + (int(i[0]) - 1) * 4 + half_day_cof).value = str(
                                end_day_hours) + ':' + str(end_day_minuts)
                    else:
                        wb['стр.1'].cell(13, 34 + (int(i[0]) - 1) * 4 + half_day_cof).value = i[1]
                if count_day_minuts == 0:
                    wb['стр.1'].cell(13, 165).value = str(count_day_hours)
                else:
                    wb['стр.1'].cell(13, 165).value = str(count_day_hours) + ':' + str(count_day_minuts)

            # ИМЯ ФАМИЛИЯ ДОЛЖНОСТЬ ID
            curr.execute("""select fname from consumer where userid = ?""", userID)
            fname = curr.fetchall()
            curr.execute("""select lname from consumer where userid = ?""", userID)
            lname = curr.fetchall()
            curr.execute("""select post from consumer where userid = ?""", userID)
            post = curr.fetchall()
            wb['стр.1'].cell(13, 1).value = fname[0][0] + ' ' + lname[0][0]
            wb['стр.1'].cell(13, 25).value = post[0][0]
            wb['стр.1'].cell(13, 13).value = userID
            # ДАТА ФОРМИРОВАНИЯ ДОКУМЕНТА
            wb['стр.1'].cell(8, 157).value = date.today()

            wb.save('db/Tabels/%sTabel.xlsx' % str(userID))
        if command == 'photographing':
            print("Enter userID")
            userID = input()
            try:
                curr.execute("SELECT userid FROM consumer WHERE userid = ?", userID)
            except:
                print('userID not found')
                continue
            capture = cv2.VideoCapture(0)
            count = 0
            while count != 15:
                ret, img = capture.read()
                cv2.imshow("From Camera", img)
                photo(img, userID, count)
                count += 1  # Итерация для записываемых кадров
                #k = cv2.waitKey(30)  # Считывания клавишы Esc для прекращения трансляции изображения
                #if k == 27:
                #    break
            capture.release()
            cv2.destroyAllWindows()
