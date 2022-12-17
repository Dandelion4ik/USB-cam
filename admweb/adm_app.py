import os
import sqlite3
from werkzeug.exceptions import abort
from flask import Flask, render_template, request, url_for, flash, redirect, send_file, render_template
from datetime import date, datetime
import openpyxl as ox
import shutil
import cv2

from admweb import admapp

UPLOAD_FOLDER = 'C:/Users/kavia/PycharmProjects/USB-cam/db/upload'
ALLOWED_EXTENSIONS = {'pdf'}
haar_face_cascade = cv2.CascadeClassifier("haarcascade_frontalface_default.xml")


def get_db_connection():
    conn = sqlite3.connect('../AllPersons.db')
    conn.row_factory = sqlite3.Row
    return conn


def get_post(post_id):
    conn = get_db_connection()
    post = conn.execute('SELECT * FROM consumer WHERE userid = ?',
                        (post_id,)).fetchone()
    conn.close()
    if post is None:
        abort(404)
    return post


def identi(img, userID, count):
    scale_factor = 1.1  # коэфицент увеличения размера окна поиска на каждой итерации
    min_neighbords = 6  # размер окна
    gray_img = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    faces = haar_face_cascade.detectMultiScale(gray_img, scale_factor, min_neighbords)  # Поиск всех лиц
    for (x, y, w, h) in faces:
        f = cv2.resize(gray_img[y:y + h, x:x + w], (200, 200))  # Создание кадра для идентицикатора
        cv2.imwrite('../db/identify/%s_%s.jpg' % (userID, str(count)), f)  # Запись этого кадра


@admapp.route('/')
def index():
    conn = get_db_connection()
    posts = conn.execute('SELECT * FROM consumer').fetchall()
    conn.close()
    return render_template('index.html', posts=posts)


@admapp.route('/<int:post_id>')
def post(post_id):
    post = get_post(post_id)
    return render_template('post.html', post=post)


@admapp.route('/create', methods=('GET', 'POST'))
def create():
    if request.method == 'POST':
        userid = request.form['userid']
        fname = request.form['fname']
        lname = request.form['lname']
        post = request.form['post']
        file = 'db/identify/%s' % userid
        status = request.form['status']
        if not userid or not fname or not lname or not status or not post:
            flash('Title is required!')
        else:
            conn = get_db_connection()
            conn.execute(
                'INSERT INTO consumer (userid, fname, lname, post, file, status) VALUES (?, ?, ?, ?, ?, ?)',
                (userid, fname, lname, post, file, status))
            conn.commit()
            conn.close()
            return redirect(url_for('index'))

    return render_template('create.html')


@admapp.route('/<int:userid>/edit', methods=('GET', 'POST'))
def edit(userid):
    post = get_post(userid)
    if request.method == 'POST':
        fname = request.form['fname']
        lname = request.form['lname']
        post = request.form['post']
        status = request.form['status']
        conn = get_db_connection()
        conn.execute('UPDATE consumer SET fname = ?, lname = ?, post = ?, status = ?'
                     ' WHERE userid = ?',
                     (fname, lname, post, status, userid))
        conn.commit()
        conn.close()
        return redirect(url_for('index'))
    return render_template('edit.html', post=post)


@admapp.route('/<int:userid>/delete', methods=('POST',))
def delete(userid):
    post = get_post(userid)
    conn = get_db_connection()
    conn.execute('DELETE FROM consumer WHERE userid = ?', (userid,))
    conn.commit()
    conn.close()
    flash('"{}" was successfully deleted!'.format(post['userid']))
    return redirect(url_for('index'))


@admapp.route('/tabel', methods=('GET', 'POST'))
def tabel():
    if request.method == 'POST':
        userid = request.form['userid']
        type = request.form['type']
        conn = get_db_connection()
        curr = conn.cursor()
        try:
            curr.execute("SELECT userid FROM consumer WHERE userid = ?", userid)
            test = curr.fetchall()
            test2 = test[0]
        except:
            flash('userID not found')
            return redirect(url_for('index'))
        print("Enter type of work")
        shutil.copyfile('Tabel.xlsx',
                        '../db/Tabels/%sTabel.xlsx' % str(userid))  # CКОПИРОВАЛИ НЕЗАПОЛНЕННЫЙ ТАБЕЛЬ В ПАПКУ ЮЗЕРА
        wb = ox.load_workbook('../db/Tabels/%sTabel.xlsx' % str(userid))
        # НОМЕР ТАБЕЛЯ
        wb['стр.1'].cell(1, 79).value = userid
        # ПЕРИОД
        wb['стр.1'].cell(4, 70).value = '30'
        wb['стр.1'].cell(4, 77).value = 'Ноября'
        wb['стр.1'].cell(4, 100).value = '22'
        # УЧРЕЖДЕНИЕ
        wb['стр.1'].cell(5, 25).value = 'МГТУ им. Баумана'
        # СТРУКТУРНОЕ ПОДРАЗДЕЛЕНИЕ
        wb['стр.1'].cell(6, 25).value = 'ИУ8-73'
        # ВИД ТАБЕЛЯ
        wb['стр.1'].cell(7, 25).value = 'Первичный'
        # СТАТИСТИКА РАБОЧИХ ДНЕЙ ИЗ БД
        if type == 'жесткий':
            curr.execute("""select dates, status from time_tracking where userid = ?""", userid)
            statistic_days = curr.fetchall()
            half_day_cof = 0
            count_working_day = 0
            for i in statistic_days:
                if i[0] == 16:
                    half_day_cof = 7
                    wb['стр.1'].cell(13, 94).value = count_working_day
                if i[1] == 'Я':
                    count_working_day += 1
                if i[1] == 'Я':
                    wb['стр.1'].cell(13, 34 + (int(i[0]) - 1) * 4 + half_day_cof).value = i[1]
                else:
                    wb['стр.1'].cell(13, 34 + (int(i[0]) - 1) * 4 + half_day_cof).value = i[1]
            wb['стр.1'].cell(13, 165).value = count_working_day
        elif type == 'свободный':
            curr.execute("""select dates, status from time_tracking where userid = ?""", userid)
            statistic_days = curr.fetchall()
            curr.execute("""select dates, time_b, time_e from schedule where userid = ?""", userid)
            schedule_days = curr.fetchall()
            half_day_cof = 0
            count_day_hours = 0
            count_day_minuts = 0
            for i in statistic_days:
                day, begin_day, end_day = schedule_days[i[0] - 1]
                end_day = datetime.strptime(end_day, '%H:%M:%S')
                begin_day = datetime.strptime(begin_day, '%H:%M:%S')
                end_day_hours = end_day.hour
                end_day_minuts = end_day.minute
                begin_day_hours = begin_day.hour
                begin_day_minuts = begin_day.minute
                if i[0] == 16:
                    half_day_cof = 7
                    if count_day_minuts == 0:
                        wb['стр.1'].cell(13, 94).value = str(count_day_hours)
                    else:
                        wb['стр.1'].cell(13, 94).value = str(count_day_hours) + ':' + str(count_day_minuts)
                if i[1] == 'Я':
                    if end_day_minuts - begin_day_minuts < 0:
                        end_day_hours -= 1
                        end_day_minuts += 60
                        end_day_minuts -= begin_day_minuts
                    else:
                        end_day_minuts -= begin_day_minuts
                    end_day_hours -= begin_day_hours
                    count_day_hours += end_day_hours
                    if count_day_minuts + end_day_minuts > 60:
                        count_day_hours += 1
                        count_day_minuts += 60
                        count_day_minuts -= end_day_minuts
                if i[1] == 'Я':
                    if end_day_minuts == 0:
                        wb['стр.1'].cell(13, 34 + (int(i[0]) - 1) * 4 + half_day_cof).value = str(
                            end_day_hours)
                    else:
                        wb['стр.1'].cell(13, 34 + (int(i[0]) - 1) * 4 + half_day_cof).value = str(
                            end_day_hours) + ':' + str(end_day_minuts)
                else:
                    wb['стр.1'].cell(13, 34 + (int(i[0]) - 1) * 4 + half_day_cof).value = i[1]
            if count_day_minuts == 0:
                wb['стр.1'].cell(13, 165).value = str(count_day_hours)
            else:
                wb['стр.1'].cell(13, 165).value = str(count_day_hours) + ':' + str(count_day_minuts)

        # ИМЯ ФАМИЛИЯ ДОЛЖНОСТЬ ID
        curr.execute("""select fname from consumer where userid = ?""", userid)
        fname = curr.fetchall()
        curr.execute("""select lname from consumer where userid = ?""", userid)
        lname = curr.fetchall()
        curr.execute("""select post from consumer where userid = ?""", userid)
        post = curr.fetchall()
        wb['стр.1'].cell(13, 1).value = fname[0][0] + ' ' + lname[0][0]
        wb['стр.1'].cell(13, 25).value = post[0][0]
        wb['стр.1'].cell(13, 13).value = userid
        # ДАТА ФОРМИРОВАНИЯ ДОКУМЕНТА
        wb['стр.1'].cell(8, 157).value = date.today()
        wb.save('../db/Tabels/%sTabel.xlsx' % str(userid))
        flash('Tabel was successfully generated!')
        return redirect(url_for('index'))
    return render_template('tabel.html')


@admapp.route('/photo', methods=('GET', 'POST'))
def photo():
    if request.method == 'POST':
        userid = request.form['userid']
        conn = get_db_connection()
        curr = conn.cursor()
        try:
            curr.execute("SELECT userid FROM consumer WHERE userid = ?", userid)
            test = curr.fetchall()
            test2 = test[0]
        except:
            flash('userID not found')
            return redirect(url_for('index'))
        capture = cv2.VideoCapture(0)
        count = 0
        while count != 15:
            ret, img = capture.read()
            cv2.imshow("From Camera", img)
            identi(img, userid, count)
            count += 1  # Итерация для записываемых кадров
            # k = cv2.waitKey(30)  # Считывания клавишы Esc для прекращения трансляции изображения
            # if k == 27:
            #    break
        capture.release()
        cv2.destroyAllWindows()
        flash('Photo was successfully generated!')
        return redirect(url_for('index'))

    return render_template('photo.html')


admapp.run('127.0.0.2', debug=True)
